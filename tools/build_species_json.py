#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
국가생물종목록(`docs/_input/생물종 일괄입력.xlsx`) → `species.json` 의 조류·포유류·양서파충류 갱신.

- 어류·저서동물은 관리자 마스터(FishSpeciesList·BenthosSpeciesList)가 기준이므로 **기존 값을 그대로 보존**한다.
- 원본 목록이 갱신되면 이 스크립트를 다시 실행한다(재실행 가능).
- 종명은 앞뒤 공백만 제거하고 원문을 변형하지 않는다(CLAUDE.md §7.3).

사용법:
    python tools/build_species_json.py
    python tools/build_species_json.py --input <xlsx> --output <species.json> --version nibr-20260820
"""
from __future__ import annotations

import argparse
import datetime as _dt
import io
import json
import re
import sys
from pathlib import Path

try:
    import openpyxl
except ImportError:  # pragma: no cover
    sys.exit("openpyxl 이 필요합니다:  pip install openpyxl")

# ── 분류군 시트의 열 위치(0-based). 헤더 텍스트로 검증한 뒤 사용한다 ──────────────
COL_KTSN = 0        # 행1 = 'KTSN'
COL_ORDER_KO = 8    # 행2 = '국명'  (Order 국명)
COL_FAMILY_KO = 10  # 행2 = '국명'  (Family 국명)
COL_SPECIES_EN = 26 # 행1 = '학명'   (완성형 — Genus+Species 조립 불필요)
COL_SPECIES_KO = 27 # 행1 = '대표국명'
DATA_START_ROW = 3  # 행1·행2가 헤더

# 우리 앱 분류군 ← 원본 시트(양서파충류는 두 시트를 합친다)
TAXON_SHEETS = {
    "Bird": ["조류"],
    "Mammal": ["포유류"],
    "Amphibian": ["양서류", "파충류"],
}

# 보호종 시트의 '분류군' 값 (분류군마다 표기가 다르다)
TAXON_GROUP_LABELS = {
    "Bird": {"조류"},
    "Mammal": {"포유류"},
    "Amphibian": {"양서류·파충류", "파충류", "양서류"},
}

ENDANGERED_SHEET = "멸종위기 야생생물 목록"
MONUMENT_SHEET = "천연기념물 목록"
INVASIVE_SHEET = "생태계교란생물 목록"


def s(v) -> str | None:
    """셀 값을 문자열로. 앞뒤 공백만 제거하고 빈 값은 None."""
    if v is None:
        return None
    t = str(v).strip()
    return t or None


def check_headers(ws, sheet_name: str) -> None:
    """열 번호가 원본과 맞는지 헤더 텍스트로 검증. 어긋나면 즉시 중단한다."""
    rows = list(ws.iter_rows(min_row=1, max_row=2, values_only=True))
    if len(rows) < 2:
        raise SystemExit(f"[{sheet_name}] 헤더 2행을 찾을 수 없습니다.")
    r1, r2 = rows[0], rows[1]

    def at(row, idx):
        return s(row[idx]) if idx < len(row) else None

    expected = [
        (COL_KTSN, r1, "KTSN", "행1"),
        (COL_SPECIES_EN, r1, "학명", "행1"),
        (COL_SPECIES_KO, r1, "대표국명", "행1"),
        (COL_ORDER_KO, r2, "국명", "행2"),
        (COL_FAMILY_KO, r2, "국명", "행2"),
    ]
    for idx, row, want, where in expected:
        got = at(row, idx)
        if got != want:
            raise SystemExit(
                f"[{sheet_name}] 열 구조가 바뀌었습니다. {where} {idx}번 열 = {got!r} (기대: {want!r})\n"
                f"→ 원본 파일이 갱신된 것 같습니다. tools/build_species_json.py 의 열 상수를 확인하세요."
            )


def read_taxon_sheet(wb, sheet_name: str, report: dict) -> list[dict]:
    """분류군 시트 1개를 읽어 종 목록으로. 대표국명이 빈 행은 건너뛴다."""
    ws = wb[sheet_name]
    check_headers(ws, sheet_name)

    out: list[dict] = []
    for row_no, row in enumerate(
        ws.iter_rows(min_row=DATA_START_ROW, values_only=True), start=DATA_START_ROW
    ):
        ko = s(row[COL_SPECIES_KO]) if COL_SPECIES_KO < len(row) else None
        if not ko:
            # 상위 분류 행 등 — 대표국명이 없으면 종이 아니다
            if any(s(c) for c in row):
                report["blank_ko"].append(f"{sheet_name} 행{row_no}")
            continue
        out.append(
            {
                "SpeciesKo": ko,
                "SpeciesEn": s(row[COL_SPECIES_EN]) if COL_SPECIES_EN < len(row) else None,
                "OrderKo": s(row[COL_ORDER_KO]) if COL_ORDER_KO < len(row) else None,
                "FamilyKo": s(row[COL_FAMILY_KO]) if COL_FAMILY_KO < len(row) else None,
                "Endangered1": None,
                "Endangered2": None,
                "NaturalMonument": None,
                "Invasive": None,
                "Ktsn": s(row[COL_KTSN]) if COL_KTSN < len(row) else None,
            }
        )
    return out


def dedupe(items: list[dict], taxon: str, report: dict) -> list[dict]:
    """국명 중복은 첫 항목만 채택하고, 건너뛴 항목을 보고한다(학명이 달라 임의 병합 금지)."""
    seen: dict[str, dict] = {}
    kept: list[dict] = []
    for it in items:
        ko = it["SpeciesKo"]
        if ko in seen:
            report["dupes"].append(
                f"{taxon}: {ko}  (채택 {seen[ko]['SpeciesEn']}  / 건너뜀 {it['SpeciesEn']})"
            )
            continue
        seen[ko] = it
        kept.append(it)
    return kept


def read_protection(wb, sheet_name: str, header_row: int, col_group: int,
                    col_ko: int, col_en: int, col_grade: int | None = None) -> list[dict]:
    """보호종 시트 1개를 읽는다. 국명이 빈 행(상위 지정 등)은 건너뛴다."""
    ws = wb[sheet_name]
    out = []
    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        def at(i):
            return s(row[i]) if i is not None and i < len(row) else None
        ko = at(col_ko)
        if not ko:
            continue
        out.append({
            "group": at(col_group),
            "ko": ko,
            "en": at(col_en),
            "grade": at(col_grade) if col_grade is not None else None,
        })
    return out


def genus_of(scientific_name: str | None) -> str | None:
    """학명에서 속 이름(첫 단어)만 뽑는다. 예: 'Trachemys scripta' → 'Trachemys'."""
    if not scientific_name:
        return None
    parts = scientific_name.split()
    return parts[0] if parts else None


def is_genus_level(scientific_name: str | None) -> bool:
    """속 전체를 가리키는 등재인지. 예: 'Trachemys spp.' / 'Rana sp.'"""
    if not scientific_name:
        return False
    return bool(re.search(r"\bspp?\.\s*$", scientific_name.strip()))


def looks_like_group_name(korean_name: str | None) -> bool:
    """국명이 종이 아니라 무리를 가리키는지. 예: '붉은귀거북속 전종', '고니류'."""
    if not korean_name:
        return False
    return "전종" in korean_name or korean_name.endswith("류")


def apply_flags(species: list[dict], entries: list[dict], taxon: str,
                assign, report: dict, label: str) -> None:
    """보호종 목록을 종 목록에 조인한다.

    ① 속 단위 등재(`Trachemys spp.`)면 **그 속의 모든 종**에 플래그를 준다.
       (이 처리가 없으면 붉은귀거북 같은 흔한 교란종이 표시되지 않는다)
    ② 그 외에는 학명 정확 일치 → 국명 정확 일치 순으로 1종에만 준다.
    """
    by_en: dict[str, dict] = {}
    by_ko: dict[str, dict] = {}
    by_genus: dict[str, list[dict]] = {}
    for sp in species:
        if sp["SpeciesEn"]:
            by_en.setdefault(sp["SpeciesEn"], sp)
        by_ko.setdefault(sp["SpeciesKo"], sp)
        g = genus_of(sp["SpeciesEn"])
        if g:
            by_genus.setdefault(g, []).append(sp)

    labels = TAXON_GROUP_LABELS[taxon]
    hit_en = hit_ko = hit_genus = miss = 0
    for e in entries:
        if e["group"] not in labels:
            continue

        targets: list[dict] = []
        kind = ""
        if is_genus_level(e["en"]):
            targets = by_genus.get(genus_of(e["en"]) or "", [])
            kind = "속"
        else:
            t = by_en.get(e["en"]) if e["en"] else None
            if t is not None:
                targets, kind = [t], "학명"
            else:
                t = by_ko.get(e["ko"])
                if t is not None:
                    targets, kind = [t], "국명"

        if not targets:
            miss += 1
            note = ""
            # 학명 없이 국명만 무리를 가리키는 경우는 자동 확장이 불가능하므로 눈에 띄게 남긴다
            if looks_like_group_name(e["ko"]) and not is_genus_level(e["en"]):
                note = "  ← 무리(속/류) 단위로 보이나 학명이 종 단위가 아니어서 자동 확장 불가"
            report["unmatched"].append(f"{taxon} · {label}: {e['ko']} ({e['en']}){note}")
            continue

        for t in targets:
            assign(t, e)

        if kind == "속":
            hit_genus += len(targets)
            for t in targets:
                report["genus"].append(
                    f"{taxon} · {label}: {e['ko']} ({e['en']}) → {t['SpeciesKo']} ({t['SpeciesEn']})"
                )
        elif kind == "학명":
            hit_en += 1
        else:
            hit_ko += 1

    report["join"].append(
        f"  {taxon:<10} {label:<12} 학명매칭 {hit_en:>3} · 국명매칭 {hit_ko:>3} · "
        f"속단위매칭 {hit_genus:>3} · 미매칭 {miss:>3}"
    )


def main() -> int:
    root = Path(__file__).resolve().parents[1]
    ap = argparse.ArgumentParser(description="국가생물종목록 → species.json (조류·포유류·양서파충류)")
    ap.add_argument("--input", default=str(root / "docs" / "_input" / "생물종 일괄입력.xlsx"))
    ap.add_argument("--output", default=str(root / "SDSM_Surveyor_App" / "species.json"))
    ap.add_argument("--version", default=None, help="기본값: nibr-YYYYMMDD (오늘)")
    args = ap.parse_args()

    src = Path(args.input)
    dst = Path(args.output)
    if not src.exists():
        return int(bool(sys.stderr.write(f"입력 파일이 없습니다: {src}\n")))
    if not dst.exists():
        return int(bool(sys.stderr.write(f"기존 species.json 이 없습니다: {dst}\n")))

    report = {"blank_ko": [], "dupes": [], "join": [], "unmatched": [], "genus": []}

    print(f"입력 : {src}")
    wb = openpyxl.load_workbook(src, read_only=True, data_only=True)

    endangered = read_protection(wb, ENDANGERED_SHEET, header_row=2,
                                 col_group=2, col_ko=3, col_en=4, col_grade=1)
    monuments = read_protection(wb, MONUMENT_SHEET, header_row=1,
                                col_group=0, col_ko=4, col_en=5)
    invasives = read_protection(wb, INVASIVE_SHEET, header_row=1,
                                col_group=1, col_ko=3, col_en=4)

    def set_endangered(sp, e):
        if e["grade"] == "Ⅰ급":
            sp["Endangered1"] = "O"
        elif e["grade"] == "Ⅱ급":
            sp["Endangered2"] = "O"

    catalog: dict[str, list[dict]] = {}
    for taxon, sheets in TAXON_SHEETS.items():
        items: list[dict] = []
        for sheet in sheets:
            part = read_taxon_sheet(wb, sheet, report)
            print(f"  {sheet:<8} {len(part):>4}종")
            items.extend(part)
        items = dedupe(items, taxon, report)
        apply_flags(items, endangered, taxon, set_endangered, report, "멸종위기")
        apply_flags(items, monuments, taxon,
                    lambda sp, e: sp.__setitem__("NaturalMonument", "O"), report, "천연기념물")
        apply_flags(items, invasives, taxon,
                    lambda sp, e: sp.__setitem__("Invasive", "O"), report, "생태계교란")
        catalog[taxon] = items
    wb.close()

    # 기존 species.json 로드 — 어류·저서동물은 그대로 보존한다
    data = json.loads(io.open(dst, encoding="utf-8-sig").read())
    fish_before, benthos_before = len(data.get("Fish", [])), len(data.get("Benthos", []))

    version = args.version or f"nibr-{_dt.date.today():%Y%m%d}"
    data["Version"] = version
    data["GeneratedAt"] = _dt.datetime.now().replace(microsecond=0).isoformat()
    for taxon, items in catalog.items():
        data[taxon] = items

    with io.open(dst, "w", encoding="utf-8", newline="\n") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
        f.write("\n")

    # ── 보고 ────────────────────────────────────────────────────────────────
    print()
    print("보호종 조인")
    for line in report["join"]:
        print(line)

    if report["genus"]:
        print(f"\n속(genus) 단위 등재로 플래그가 붙은 종 ({len(report['genus'])}건)")
        for g in report["genus"]:
            print(f"  {g}")

    if report["dupes"]:
        print(f"\n국명 중복 — 첫 항목만 채택 ({len(report['dupes'])}건)")
        for d in report["dupes"]:
            print(f"  {d}")

    if report["blank_ko"]:
        print(f"\n대표국명이 비어 건너뛴 행 ({len(report['blank_ko'])}건)")
        for b in report["blank_ko"]:
            print(f"  {b}")

    if report["unmatched"]:
        print(f"\n보호종 목록에 있으나 종목록에서 못 찾음 ({len(report['unmatched'])}건)")
        for u in report["unmatched"]:
            print(f"  {u}")

    print()
    print(f"결과 : {dst}")
    print(f"  Version   {version}")
    print(f"  Bird      {len(catalog['Bird']):>4}종")
    print(f"  Mammal    {len(catalog['Mammal']):>4}종")
    print(f"  Amphibian {len(catalog['Amphibian']):>4}종 (양서류+파충류)")
    print(f"  Fish      {fish_before:>4}종  (보존)")
    print(f"  Benthos   {benthos_before:>4}종  (보존)")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
