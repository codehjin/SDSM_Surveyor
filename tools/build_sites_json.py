#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
제출 엑셀의 `조사지점` 시트 → 지점 마스터 `sites.json` 생성 (04_FEATURE_SITE_SESSION §A-2).

⚠ **파일마다 `조사지점` 시트의 열 구성이 다르다.** (하천별·분류군별로 제각각)
  - 곡교천 : 지점설명 | DB상 지점명 | 위도(DMS) | 경도(DMS) | 22~25 | 26 | (St번호) | 비고
  - 오산천 : 조사지점 | 지점명 | 하천차수 | 위도(십진) | 경도(십진) | 연도별 지도번호…
  - 원천리천: 지점 | DB 지점명 | 위치 | 위도 | 경도 | 연도별 표기…
  그래서 **열 번호를 고정하지 않고 행2의 헤더 텍스트로 위치를 찾는다.**

사용법:
    python tools/build_sites_json.py
    python tools/build_sites_json.py --input <폴더> --output <sites.json> --version sites-20260820
"""
from __future__ import annotations

import argparse
import datetime as _dt
import glob
import io
import json
import os
import re
import sys

try:
    import openpyxl
except ImportError:  # pragma: no cover
    sys.exit("openpyxl 이 필요합니다:  pip install openpyxl")

SHEET = "조사지점"
HEADER_ROW = 2      # 1-based. 데이터는 행3부터
DATA_ROW = 3

# 관리자 DB `SiteDivision` 기준 (하천 → 사업장)
WORKPLACE_BY_RIVER = {"오산천": "기흥", "원천리천": "화성", "곡교천": "온양"}

# 곡교천은 한 과업대표하천 안에 지점 그룹이 여러 개다(SiteDivision.WorkplaceSite)
SITE_GROUPS = {"곡교천": ["곡교천", "천안천", "회룡천", "공수리"]}

TAXA = ["어류", "저서동물", "조류", "포유류", "양서파충류", "수질", "서식수변"]

_YEAR_HEADER = re.compile(r"^\s*'?\d{2,4}\s*(~\s*\d{2,4})?\s*$")
_ST_VALUE = re.compile(r"^\s*st\.?\s*\d+", re.IGNORECASE)
_DMS = re.compile(r"""^\s*(\d+)\s*°\s*(\d+)\s*'\s*([\d.]+)\s*"?\s*([NSEW])?\s*$""")


def s(v) -> str | None:
    if v is None:
        return None
    t = str(v).strip()
    return t or None


def to_decimal(v) -> float | None:
    """위/경도를 십진수로. 엑셀에 DMS(36°45'47.09"N)와 십진수가 섞여 있다."""
    t = s(v)
    if t is None:
        return None
    m = _DMS.match(t)
    if m:
        deg, minute, sec, hemi = m.groups()
        val = int(deg) + int(minute) / 60 + float(sec) / 3600
        if hemi and hemi.upper() in ("S", "W"):
            val = -val
        return round(val, 6)
    try:
        return round(float(t), 6)
    except ValueError:
        return None


def parse_filename(path: str) -> tuple[str | None, str | None, str | None]:
    """파일명 → (대분류, 과업대표하천, 분류군). 예: 2602_방류하천_곡교천_어류_DB_v1.xlsx"""
    name = os.path.basename(path)
    project = "방류하천" if "방류하천" in name else ("생태현황" if "생태현황" in name else None)
    river = next((r for r in WORKPLACE_BY_RIVER if r in name), None)
    taxon = next((t for t in TAXA if t in name), None)
    return project, river, taxon


def find_columns(header: list) -> dict:
    """행2 헤더 텍스트로 열 위치를 찾는다(파일마다 구성이 달라 고정할 수 없다)."""
    cols: dict = {"years": {}}
    for i, raw in enumerate(header):
        h = s(raw)
        if not h:
            continue
        if "지점명" in h:
            cols.setdefault("name", i)
        elif h in ("지점설명", "조사지점", "지점"):
            cols.setdefault("desc", i)
        elif h in ("위치",):
            cols.setdefault("desc2", i)
        elif h == "위도":
            cols["lat"] = i
        elif h == "경도":
            cols["lng"] = i
        elif "하천차수" in h:
            cols["chasu"] = i
        elif h == "비고":
            cols["note"] = i
        elif _YEAR_HEADER.match(h):
            cols["years"][h.strip().lstrip("'")] = i
    # 지점명 헤더가 없으면 관례상 C열
    cols.setdefault("name", 2)
    cols.setdefault("desc", 1)
    return cols


def find_st_column(ws, cols: dict) -> int | None:
    """`St.1` 형태의 조사장소 번호 열은 헤더가 비어 있는 경우가 있어 값으로 찾는다."""
    counts: dict[int, int] = {}
    for row in ws.iter_rows(min_row=DATA_ROW, values_only=True):
        # 데이터 블록(연속 구간) 안에서만 찾는다 — 아래쪽 메모의 "※ st.3 …"에 걸리지 않도록
        head = s(row[cols["name"]]) if cols["name"] < len(row) else None
        if not head:
            break
        for i, v in enumerate(row):
            t = s(v)
            if t and _ST_VALUE.match(t):
                counts[i] = counts.get(i, 0) + 1
    if not counts:
        return None
    return max(counts, key=lambda k: counts[k])


def normalize_st(v: str) -> str:
    """`st.3` `ST 3` → `St.3` 로 표기 통일."""
    m = re.search(r"\d+", v)
    return f"St.{m.group()}" if m else v


def site_group(river: str | None, name: str) -> str | None:
    """곡교천처럼 한 하천에 지점 그룹이 여럿이면 지점명 접두어로 판별(SiteDivision.WorkplaceSite)."""
    if not river:
        return None
    for g in SITE_GROUPS.get(river, [river]):
        if name.startswith(g):
            return g
    return river


def main() -> int:
    root = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    ap = argparse.ArgumentParser(description="제출 엑셀 조사지점 시트 → sites.json")
    ap.add_argument("--input", default=r"F:\17_한국환경지리연구소\2_삼성DS\2026년\5_보낸파일\20260305\일괄입력자료")
    ap.add_argument("--output", default=os.path.join(root, "SDSM_Surveyor_App", "sites.json"))
    ap.add_argument("--version", default=None, help="기본값: sites-YYYYMMDD")
    args = ap.parse_args()

    files = sorted(glob.glob(os.path.join(args.input, "*.xlsx")))
    if not files:
        return int(bool(sys.stderr.write(f"입력 폴더에 xlsx가 없습니다: {args.input}\n")))

    merged: dict[tuple, dict] = {}
    skipped: list[str] = []
    per_file: list[str] = []

    for path in files:
        project, river, taxon = parse_filename(path)
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        try:
            if SHEET not in wb.sheetnames:
                skipped.append(f"{os.path.basename(path)} — '{SHEET}' 시트 없음")
                continue
            ws = wb[SHEET]
            rows = list(ws.iter_rows(min_row=1, max_row=HEADER_ROW, values_only=True))
            if len(rows) < HEADER_ROW:
                skipped.append(f"{os.path.basename(path)} — 헤더 행 없음")
                continue
            cols = find_columns(list(rows[HEADER_ROW - 1]))
            st_col = find_st_column(ws, cols)

            added = 0
            for row in ws.iter_rows(min_row=DATA_ROW, values_only=True):
                def at(idx):
                    return s(row[idx]) if idx is not None and idx < len(row) else None

                name = at(cols["name"])
                # ⚠ 지점 표는 행3부터 **연속**이고, 빈 행 아래에는 범례·메모가 섞여 있다
                #   (예: "2011", "2024 조사지점", "※ st.3 대황교 지점 잘못 설명됨").
                #   그래서 첫 빈 지점명에서 멈춘다. 계속 훑으면 메모가 지점으로 들어온다.
                if not name:
                    break
                if name in ("지점명", "DB상 지점명", "DB 지점명"):
                    continue

                key = (project, river, name)
                rec = merged.setdefault(key, {
                    "Project": project,
                    "Workplace": WORKPLACE_BY_RIVER.get(river or ""),
                    "River": river,
                    "SiteGroup": site_group(river, name),
                    "SiteName": name,
                    "StNo": None,
                    "Desc": None,
                    "Chasu": None,
                    "Lat": None,
                    "Lng": None,
                    "YearAliases": {},
                    "MapNumbers": {},
                    "Taxa": [],
                })

                # 원천리천은 '지점' 열이 번호(0,1,2…)라 설명이 못 된다 → '위치' 열을 우선
                d1, d2 = at(cols.get("desc")), at(cols.get("desc2"))
                if d1 and d1.replace(".", "").isdigit():
                    d1, d2 = d2, d1
                rec["Desc"] = rec["Desc"] or d1 or d2
                rec["Lat"] = rec["Lat"] if rec["Lat"] is not None else to_decimal(row[cols["lat"]] if "lat" in cols and cols["lat"] < len(row) else None)
                rec["Lng"] = rec["Lng"] if rec["Lng"] is not None else to_decimal(row[cols["lng"]] if "lng" in cols and cols["lng"] < len(row) else None)
                if rec["Chasu"] is None and "chasu" in cols:
                    c = at(cols["chasu"])
                    if c and c.isdigit():
                        rec["Chasu"] = int(c)
                if rec["StNo"] is None and st_col is not None:
                    v = at(st_col)
                    if v and _ST_VALUE.match(v):
                        rec["StNo"] = normalize_st(v)
                # 연도별 열에는 두 종류가 섞여 있다 : 그 시기의 **지점명**(오산천5)과 **지도 번호**(3).
                # 의미가 다르므로 분리해 담는다.
                for label, idx in cols["years"].items():
                    v = at(idx)
                    if not v:
                        continue
                    if v.isdigit():
                        rec["MapNumbers"].setdefault(label, v)
                    elif _ST_VALUE.match(v):
                        rec["StNo"] = rec["StNo"] or normalize_st(v)
                    else:
                        rec["YearAliases"].setdefault(label, v)

                note = at(cols.get("note"))
                if note and "수질만" in note.replace(" ", ""):
                    rec["Taxa"] = ["수질"]          # 비고가 명시하면 그것이 우선
                elif taxon and taxon not in rec["Taxa"] and rec["Taxa"] != ["수질"]:
                    rec["Taxa"].append(taxon)
                added += 1
            per_file.append(f"  {os.path.basename(path):<44} {added:>3}행  열={{name:{cols['name']}, lat:{cols.get('lat')}, lng:{cols.get('lng')}, st:{st_col}}}")
        finally:
            wb.close()

    sites = sorted(merged.values(), key=lambda r: (r["Project"] or "", r["River"] or "", r["SiteName"]))
    version = args.version or f"sites-{_dt.date.today():%Y%m%d}"
    data = {
        "Version": version,
        "GeneratedAt": _dt.datetime.now().replace(microsecond=0).isoformat(),
        "Sites": sites,
    }
    with io.open(args.output, "w", encoding="utf-8", newline="\n") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
        f.write("\n")

    print("파일별 처리")
    for line in per_file:
        print(line)
    if skipped:
        print(f"\n건너뛴 파일 ({len(skipped)})")
        for x in skipped:
            print(f"  {x}")

    no_coord = [x["SiteName"] for x in sites if x["Lat"] is None or x["Lng"] is None]
    no_st = [x["SiteName"] for x in sites if not x["StNo"]]
    print(f"\n결과 : {args.output}")
    print(f"  Version   {version}")
    print(f"  지점 수    {len(sites)}")
    for river in sorted({(x['River'] or '(없음)') for x in sites}):
        n = sum(1 for x in sites if (x["River"] or "(없음)") == river)
        print(f"    {river:<8} {n:>3}")
    print(f"  좌표 없음  {len(no_coord)}건" + (f" : {', '.join(no_coord[:10])}" if no_coord else ""))
    print(f"  St번호 없음 {len(no_st)}건" + (f" : {', '.join(no_st[:10])}…" if no_st else ""))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
