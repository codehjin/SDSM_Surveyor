# -*- coding: utf-8 -*-
"""
회귀 기준 엑셀 셀 단위 대조 (05_REFACTORING §0-2 · R0)

사용법:
    python tools\\baseline_diff.py <기준폴더> <비교폴더>
    python tools\\baseline_diff.py docs\\_baseline  C:\\Temp\\after

xlsx 는 zip 안에 생성 시각이 들어가 바이트 비교가 불가능하다. 그래서 **셀 값**만 비교한다.
파일 목록·시트 목록·시트 크기·셀 값이 모두 같아야 통과다.

종료 코드 0 = 전부 같음, 1 = 차이 있음.
"""
import io
import os
import sys
import glob

try:
    import openpyxl
except ImportError:
    print("openpyxl 이 필요합니다:  pip install openpyxl")
    sys.exit(2)

sys.stdout.reconfigure(encoding="utf-8")

# Telerik 평가판이 붙이는 시트. 내용이 실행마다 달라질 수 있어 비교에서 뺀다.
SKIP_SHEETS = {"License"}

MAX_REPORT_PER_FILE = 15


def cells(ws):
    """시트를 (행, 열) → 값 사전으로. 빈 셀은 담지 않는다."""
    out = {}
    for row in ws.iter_rows():
        for c in row:
            if c.value is not None:
                out[(c.row, c.column)] = c.value
    return out


def col_name(n):
    s = ""
    while n > 0:
        n, r = divmod(n - 1, 26)
        s = chr(65 + r) + s
    return s


def compare_file(a_path, b_path):
    """파일 하나를 비교해 차이 목록을 돌려준다."""
    diffs = []
    wa = openpyxl.load_workbook(a_path, data_only=False)
    wb = openpyxl.load_workbook(b_path, data_only=False)

    sa = [n for n in wa.sheetnames if n not in SKIP_SHEETS]
    sb = [n for n in wb.sheetnames if n not in SKIP_SHEETS]
    if sa != sb:
        diffs.append(f"시트 목록이 다름: 기준={sa} / 비교={sb}")
        return diffs

    for name in sa:
        ca, cb = cells(wa[name]), cells(wb[name])
        keys = sorted(set(ca) | set(cb))
        for k in keys:
            va, vb = ca.get(k), cb.get(k)
            if va != vb:
                r, c = k
                diffs.append(f"[{name}] {col_name(c)}{r}  기준={va!r}  비교={vb!r}")
    wa.close()
    wb.close()
    return diffs


def main():
    if len(sys.argv) < 3:
        print(__doc__)
        return 2

    base, other = sys.argv[1], sys.argv[2]
    for d in (base, other):
        if not os.path.isdir(d):
            print(f"폴더가 없습니다: {d}")
            return 2

    fa = sorted(os.path.basename(p) for p in glob.glob(os.path.join(base, "*.xlsx")))
    fb = sorted(os.path.basename(p) for p in glob.glob(os.path.join(other, "*.xlsx")))

    print(f"기준 : {base}  ({len(fa)}개)")
    print(f"비교 : {other}  ({len(fb)}개)\n")

    bad = 0
    only_a = [f for f in fa if f not in fb]
    only_b = [f for f in fb if f not in fa]
    if only_a:
        print(f"  ! 비교 폴더에 없음: {only_a}")
        bad += len(only_a)
    if only_b:
        print(f"  ! 기준 폴더에 없음: {only_b}")
        bad += len(only_b)

    for f in fa:
        if f not in fb:
            continue
        diffs = compare_file(os.path.join(base, f), os.path.join(other, f))
        if not diffs:
            print(f"  같음  {f}")
        else:
            bad += 1
            print(f"  다름  {f}  — 차이 {len(diffs)}건")
            for d in diffs[:MAX_REPORT_PER_FILE]:
                print(f"          {d}")
            if len(diffs) > MAX_REPORT_PER_FILE:
                print(f"          … 외 {len(diffs) - MAX_REPORT_PER_FILE}건")

    print()
    if bad == 0:
        print(f"===== 셀 값 전부 일치 ({len(fa)}개 파일) =====")
        return 0
    print(f"===== 차이 있는 파일 {bad}개 =====")
    return 1


if __name__ == "__main__":
    sys.exit(main())
